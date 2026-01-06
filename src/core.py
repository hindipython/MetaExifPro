from abc import ABC, abstractmethod
import os
import datetime
from typing import Dict, Any, Optional

import mutagen
from mutagen.easyid3 import EasyID3
import piexif
from PIL import Image, ExifTags
import pypdf
from docx import Document as DocxDocument
from openpyxl import load_workbook

# Helper to reverse ExifTags for saving (Name -> ID)
TAG_NAME_TO_ID = {v: k for k, v in ExifTags.TAGS.items()}

class FileHandler(ABC):
    @abstractmethod
    def load(self, path: str) -> Dict[str, str]:
        pass

    @abstractmethod
    def save(self, path: str, data: Dict[str, str]) -> None:
        pass

class AudioHandler(FileHandler):
    """Handles Audio/Video via Mutagen (MP3, MP4, FLAC, etc). Returns ALL raw tags."""
    def load(self, path: str) -> Dict[str, str]:
        data = {}
        try:
            # mutagen.File covers MP3, MP4, FLAC, OGG, etc.
            # We do NOT use easy=True to get raw tags.
            audio = mutagen.File(path)
            if audio is None: return {}
            
            # Helper to stringify values (Mutagen values can be lists, bytes, or objects)
            def fmt(v):
                if isinstance(v, list) or isinstance(v, tuple):
                    return "; ".join([str(x) for x in v])
                return str(v)

            if hasattr(audio, "tags") and audio.tags:
                for k, v in audio.tags.items():
                    data[str(k)] = fmt(v)
            else:
                # Some formats act as dictionary directly
                for k, v in audio.items():
                     data[str(k)] = fmt(v)

            # Add stream info (Read-only usually)
            if audio.info:
                if hasattr(audio.info, 'length'): 
                    m, s = divmod(audio.info.length, 60)
                    data['@Duration'] = f"{int(m):02d}:{int(s):02d} ({audio.info.length:.2f}s)"
                if hasattr(audio.info, 'bitrate'): 
                    data['@Bitrate'] = f"{int(audio.info.bitrate / 1000)} kbps"
                if hasattr(audio.info, 'sample_rate'):
                    data['@SampleRate'] = f"{audio.info.sample_rate} Hz"
                if hasattr(audio.info, 'channels'):
                    data['@Channels'] = str(audio.info.channels)
                if hasattr(audio.info, 'encoder_info'):
                    data['@Encoder'] = str(audio.info.encoder_info)

            return data
        except Exception as e:
            print(f"Audio load error: {e}")
            return {}

    def save(self, path: str, data: Dict[str, str]) -> None:
        try:
            audio = mutagen.File(path)
            if audio is None: return 
            if audio.tags is None: audio.add_tags()
            
            for k, v in data.items():
                if k.startswith("@"): continue # Skip read-only props
                try:
                    audio.tags[k] = [v]
                except:
                    try: audio.tags[k] = v
                    except: pass
            
            current_keys = list(audio.tags.keys())
            for k in current_keys:
                if k not in data and not k.startswith("@"):
                     diff_k = str(k)
                     if diff_k not in data:
                        del audio.tags[k]
            audio.save()
        except Exception as e:
            print(f"Audio save error: {e}")

class ImageHandler(FileHandler):
    """Handles Images. aggressively reads Exif and generic Info."""
    def load(self, path: str) -> Dict[str, str]:
        data = {}
        try:
            with Image.open(path) as img:
                # Basic Image Properties
                data["@Resolution"] = f"{img.width}x{img.height}"
                data["@Format"] = str(img.format)
                data["@Mode"] = str(img.mode)
                if hasattr(img, "n_frames") and img.n_frames > 1:
                    data["@Frames"] = str(img.n_frames)

                img.load()
                
                # 1. Standard Exif
                exif = img.getexif()
                if exif:
                    for k, v in exif.items():
                         key_name = ExifTags.TAGS.get(k, f"Exif:{k}")
                         if isinstance(v, bytes):
                             try: v = v.decode().strip('\x00')
                             except: v = str(v)
                         data[key_name] = str(v)

                # 2. Deep Dive (GPS, Interop)
                if "exif" in img.info:
                    try:
                        exif_dict = piexif.load(img.info["exif"])
                        for ifd in ["0th", "Exif", "GPS", "1st", "Interop"]:
                            if ifd in exif_dict:
                                for tag_id, val in exif_dict[ifd].items():
                                    tag_name = ExifTags.TAGS.get(tag_id, f"Unknown_{tag_id}")
                                    key = f"{ifd}:{tag_name}"
                                    if tag_name in data: continue 
                                    if isinstance(val, bytes):
                                        try: val = val.decode().strip('\x00')
                                        except: val = f"<Binary {len(val)}>"
                                    data[key] = str(val)
                    except: pass

                # 3. Info Dict
                for k, v in img.info.items():
                    if k in ['exif']: continue
                    if isinstance(v, (str, int, float)):
                        data[f"Info:{k}"] = str(v)
                    elif isinstance(v, bytes):
                        try: data[f"Info:{k}"] = v.decode()
                        except: data[f"Info:{k}"] = f"<Binary {len(v)} bytes>"
                            
            return data
        except Exception as e:
            print(f"Image load error: {e}")
            return {}

    def save(self, path: str, data: Dict[str, str]) -> None:
        try:
            print(f"\n{'='*50}")
            print(f"[ImageHandler] SAVING TO: {path}")
            print(f"[ImageHandler] Total keys to process: {len(data)}")
            
            # Step 1: Load current exif (if exists)
            img = Image.open(path)
            img_format = img.format or "JPEG"
            print(f"[ImageHandler] Format: {img_format}")
            
            try: 
                exif_dict = piexif.load(img.info.get("exif", b""))
            except: 
                exif_dict = {"0th": {}, "Exif": {}, "GPS": {}, "1st": {}, "thumbnail": None}
            
            # Save image to memory buffer and CLOSE file handle
            from io import BytesIO
            img_bytes = BytesIO()
            if img_format.upper() in ["JPEG", "JPG"]:
                img.save(img_bytes, format="JPEG", quality=95)
            else:
                img.save(img_bytes, format=img_format)
            img.close()
            
            # Step 2: Modify exif_dict
            tags_written = 0
            for key, val in data.items():
                if key.startswith("@") or key.startswith("Info:") or key.startswith("File:"): 
                    continue
                
                # Skip technical tags that cause piexif errors (they're auto-managed)
                SKIP_TAGS = ["XResolution", "YResolution", "ResolutionUnit", "YCbCrPositioning", 
                             "ExifOffset", "ComponentsConfiguration", "FlashPixVersion", "Compression",
                             "JPEGInterchangeFormat", "JPEGInterchangeFormatLength"]
                tag_base = key.split(":")[-1] if ":" in key else key
                if tag_base in SKIP_TAGS:
                    continue

                tag_id = None
                target_ifd = None
                
                # Handle prefixed keys like "0th:Make" or "Exif:DateTimeOriginal"
                if ":" in key:
                    parts = key.split(":", 1)
                    prefix = parts[0]
                    tag_name = parts[1] if len(parts) > 1 else key
                    
                    # Check if prefix is IFD name
                    if prefix in ["0th", "Exif", "GPS", "1st", "Interop"]:
                        target_ifd = prefix
                        tag_id = TAG_NAME_TO_ID.get(tag_name)
                    else:
                        # Try full key or just the part after colon
                        tag_id = TAG_NAME_TO_ID.get(key) or TAG_NAME_TO_ID.get(tag_name)
                else:
                    tag_id = TAG_NAME_TO_ID.get(key)
                
                if tag_id is None:
                    print(f"  [SKIP] '{key}' -> No tag ID found")
                    continue

                # Type casting for different tag types
                INT_TAGS = [34855, 34850, 41987, 41986, 37383, 37385, 40961, 274, 41989]  # ISO, ExposureProgram, FocalLength35mm, etc
                RATIONAL_TAGS = [37377, 37378, 37379, 37380, 37381, 37386, 33434, 33437, 282, 283]  # Shutter, Aperture, Focal, XRes, YRes
                
                val_final = val
                
                if tag_id in INT_TAGS:
                    try: 
                        val_final = int(val)
                    except: 
                        pass
                        
                elif tag_id in RATIONAL_TAGS:
                    # Parse rational: "1/100" -> (1, 100), "1.7" -> (17, 10), "23" -> (23, 1)
                    # Also handle already-tuple strings: "(190, 100)"
                    try:
                        val_str = str(val).strip()
                        
                        # Already a tuple string like "(190, 100)"
                        if val_str.startswith("(") and val_str.endswith(")"):
                            inner = val_str[1:-1]
                            parts = inner.split(",")
                            val_final = (int(parts[0].strip()), int(parts[1].strip()))
                        elif "/" in val_str:
                            parts = val_str.split("/")
                            val_final = (int(parts[0]), int(parts[1]))
                        elif "." in val_str:
                            f = float(val_str)
                            val_final = (int(f * 100), 100)
                        else:
                            val_final = (int(val_str), 1)
                    except Exception as e:
                        print(f"    -> Rational parse failed: {e}")
                        val_final = (0, 1)  # Default safe value
                
                # Encode string values to bytes
                if isinstance(val_final, str):
                    val_encoded = val_final.encode('utf-8')
                elif isinstance(val_final, tuple):
                    val_encoded = val_final  # Keep as tuple for piexif
                else:
                    val_encoded = val_final  # int or other

                # Place into correct IFD
                placed = False
                
                # If we have a target IFD from prefix, use it
                if target_ifd and target_ifd in exif_dict:
                    exif_dict[target_ifd][tag_id] = val_encoded
                    placed = True
                    print(f"  [WRITE] '{key}' -> ID {tag_id} -> IFD '{target_ifd}' = '{val}'")
                else:
                    # Check existing IFDs
                    for ifd in ["0th", "Exif", "GPS", "1st"]:
                        if ifd in exif_dict and tag_id in exif_dict[ifd]:
                            exif_dict[ifd][tag_id] = val_encoded
                            placed = True
                            print(f"  [UPDATE] '{key}' -> ID {tag_id} -> IFD '{ifd}' = '{val}'")
                            break
                    
                    if not placed:
                        # Default placement
                        if tag_id in [271, 272, 305, 306, 315]:  # Make, Model, Software, DateTime, Artist
                            exif_dict["0th"][tag_id] = val_encoded
                            print(f"  [NEW->0th] '{key}' -> ID {tag_id} = '{val}'")
                        else:
                            exif_dict["Exif"][tag_id] = val_encoded
                            print(f"  [NEW->Exif] '{key}' -> ID {tag_id} = '{val}'")
                
                tags_written += 1

            print(f"[ImageHandler] Tags written: {tags_written}")
            
            # Step 3: Dump new exif and save
            exif_bytes = piexif.dump(exif_dict)
            print(f"[ImageHandler] Exif bytes size: {len(exif_bytes)}")
            
            # Re-open from memory buffer and save with new exif
            img_bytes.seek(0)
            final_img = Image.open(img_bytes)
            
            # Save to temp file first
            temp_path = path + ".tmp"
            if img_format.upper() in ["JPEG", "JPG"]:
                final_img.save(temp_path, format="JPEG", exif=exif_bytes, quality=95)
            else:
                if img_format.upper() == "PNG":
                    print("[ImageHandler] WARNING: PNG does not support EXIF!")
                    final_img.save(temp_path, format="PNG")
                else:
                    final_img.save(temp_path, format=img_format, exif=exif_bytes)
            
            final_img.close()
            
            # Replace original with temp
            os.replace(temp_path, path)
            print(f"[ImageHandler] SUCCESS! File saved.")
            print(f"{'='*50}\n")
            
        except Exception as e:
            print(f"[ImageHandler] ERROR: {e}")
            import traceback
            traceback.print_exc()


class PDFHandler(FileHandler):
    def load(self, path: str) -> Dict[str, str]:
        try:
            reader = pypdf.PdfReader(path)
            data = {}
            # Pages
            data["@Pages"] = str(len(reader.pages))
            
            meta = reader.metadata
            if meta:
                for k, v in meta.items():
                    data[k.lstrip('/')] = str(v)
            return data
        except Exception:
            return {}

    def save(self, path: str, data: Dict[str, str]) -> None:
        try:
            reader = pypdf.PdfReader(path)
            writer = pypdf.PdfWriter()
            writer.append_pages_from_reader(reader)
            
            meta_args = {f"/{k}": v for k, v in data.items() if not k.startswith("@")}
            writer.add_metadata(meta_args)
            
            temp = path + ".tmp"
            with open(temp, "wb") as f:
                writer.write(f)
            os.replace(temp, path)
        except Exception as e:
            print(f"PDF save error: {e}")

class DocxHandler(FileHandler):
    def load(self, path: str) -> Dict[str, str]:
        try:
            doc = DocxDocument(path)
            props = doc.core_properties
            data = {}
            for prop in dir(props):
                if not prop.startswith('_') and not callable(getattr(props, prop)):
                    val = getattr(props, prop)
                    if val and isinstance(val, (str, int, float)):
                        data[prop] = str(val)
            return data
        except Exception:
            return {}

    def save(self, path: str, data: Dict[str, str]) -> None:
        try:
            doc = DocxDocument(path)
            props = doc.core_properties
            for k, v in data.items():
                if hasattr(props, k):
                    try: setattr(props, k, v)
                    except: pass
            doc.save(path)
        except Exception as e:
            print(f"DOCX save error: {e}")

class XlsxHandler(FileHandler):
    def load(self, path: str) -> Dict[str, str]:
        try:
            wb = load_workbook(path)
            props = wb.properties
            data = {}
            for prop in dir(props):
                if not prop.startswith('_') and not callable(getattr(props, prop)):
                    val = getattr(props, prop)
                    if val and isinstance(val, (str, int, float)):
                        data[prop] = str(val)
            return data
        except Exception:
            return {}

    def save(self, path: str, data: Dict[str, str]) -> None:
        try:
            wb = load_workbook(path)
            props = wb.properties
            for k, v in data.items():
                 if hasattr(props, k):
                    try: setattr(props, k, v)
                    except: pass
            wb.save(path)
        except Exception as e:
            print(f"XLSX save error: {e}")

class GenericHandler(FileHandler):
    """Handles any file type just for file system stats (Dates)."""
    def load(self, path: str) -> Dict[str, str]:
        return {} # No internal metadata
    def save(self, path: str, data: Dict[str, str]) -> None:
        pass # No internal metadata to save

class MetadataManager:
    HANDLERS = {
        # Audio / Video
        ".mp3": AudioHandler(), ".flac": AudioHandler(), ".ogg": AudioHandler(), ".m4a": AudioHandler(),
        ".wav": AudioHandler(), ".wma": AudioHandler(), ".aac": AudioHandler(), ".aiff": AudioHandler(),
        ".opus": AudioHandler(),
        
        # Video Containers (Mutagen supports some)
        ".mp4": AudioHandler(), ".mov": AudioHandler(), ".mkv": AudioHandler(), ".webm": AudioHandler(),
        
        # Images
        ".jpg": ImageHandler(), ".jpeg": ImageHandler(), ".tiff": ImageHandler(), ".webp": ImageHandler(),
        ".png": ImageHandler(), ".bmp": ImageHandler(), ".gif": ImageHandler(), ".ico": ImageHandler(),
        ".heic": ImageHandler(), # Requires pillow-heif usually, but we register it just in case user has it
        
        # Documents
        ".pdf": PDFHandler(),
        ".docx": DocxHandler(), ".xlsx": XlsxHandler(),
        
        # Generic (Text/Archives code handled by generic)
        ".txt": GenericHandler(), ".md": GenericHandler(), ".csv": GenericHandler(),
        ".zip": GenericHandler(), ".rar": GenericHandler(), ".7z": GenericHandler(),
        ".exe": GenericHandler(), ".dll": GenericHandler(),
        ".py": GenericHandler(), ".js": GenericHandler(), ".html": GenericHandler(),
        ".css": GenericHandler(), ".json": GenericHandler(), ".xml": GenericHandler()
    }

    @staticmethod
    def get_handler(filepath: str) -> Optional[FileHandler]:
        _, ext = os.path.splitext(filepath)
        handler = MetadataManager.HANDLERS.get(ext.lower())
        if handler is None:
            # Fallback to generic for any unknown file to allow Date Editing
            return GenericHandler()
        return handler

    @staticmethod
    def load(filepath: str) -> Dict[str, str]:
        handler = MetadataManager.get_handler(filepath)
        data = {}
        
        # 1. Load Format-Specific Tags
        if handler:
            data = handler.load(filepath)
        
        # 2. Add Generic File System Stats (Like ExifTool)
        try:
            stat = os.stat(filepath)
            data["File:Size"] = f"{stat.st_size / 1024:.2f} KB"
            data["File:Modified"] = datetime.datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
            data["File:Created"] = datetime.datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S')
            data["File:Path"] = filepath
        except Exception:
            pass

        return data




    @staticmethod
    def save(filepath: str, data: Dict[str, str]):
        handler = MetadataManager.get_handler(filepath)
        if handler:
            # 1. Save Internal Metadata (Exif, etc)
            writable_data = {k: v for k, v in data.items() if not k.startswith("@") and not k.startswith("File:")}
            handler.save(filepath, writable_data)
            
            # 2. Sync File System Dates
            # Priority: Exif dates > File:Created/Modified (for mimicry, EXIF dates are authoritative)
            exif_date = None
            for key in ["Exif:DateTimeOriginal", "0th:DateTime", "DateTime"]:
                if key in data and data[key]:
                    exif_date = data[key]
                    break
            
            # Use EXIF date if available, otherwise fall back to File: dates
            created_date = exif_date or data.get("File:Created")
            modified_date = exif_date or data.get("File:Modified")
            
            if created_date or modified_date:
                MetadataManager.set_file_dates(filepath, created_date, modified_date)
                print(f"[MetadataManager] File dates set: Created={created_date}, Modified={modified_date}")

    @staticmethod
    def set_file_dates(filepath: str, created_str: str = None, modified_str: str = None):
        """Sets file creation and modification times on Windows."""
        formats = [
            "%Y-%m-%d %H:%M:%S",
            "%Y:%m:%d %H:%M:%S",
            "%Y/%m/%d %H:%M:%S",
            "%d.%m.%Y %H:%M:%S",
            "%Y-%m-%d %H-%M-%S",
        ]
        
        def parse_date(date_str):
            if not date_str: return None
            for fmt in formats:
                try:
                    return datetime.datetime.strptime(str(date_str).strip(), fmt).timestamp()
                except ValueError:
                    continue
            print(f"[set_file_dates] Could not parse: {date_str}")
            return None
        
        created_ts = parse_date(created_str)
        modified_ts = parse_date(modified_str)
        
        if not created_ts and not modified_ts:
            return
        
        try:
            # 1. Set Modified/Access Time via os.utime
            if modified_ts:
                os.utime(filepath, (modified_ts, modified_ts))
            
            # 2. Set Creation Time (Windows Only) via kernel32
            if os.name == 'nt' and (created_ts or modified_ts):
                import ctypes
                
                FILE_WRITE_ATTRIBUTES = 0x0100
                OPEN_EXISTING = 3
                kernel32 = ctypes.WinDLL('kernel32', use_last_error=True)
                
                h = kernel32.CreateFileW(
                    filepath, 
                    FILE_WRITE_ATTRIBUTES, 
                    0, None, OPEN_EXISTING, 
                    128,  # FILE_ATTRIBUTE_NORMAL
                    None
                )
                
                if h != -1:
                    def ts_to_filetime(ts):
                        if ts is None: return None
                        intervals = int((ts + 11644473600) * 10000000)
                        return ctypes.c_int64(intervals)
                    
                    ft_created = ts_to_filetime(created_ts)
                    ft_modified = ts_to_filetime(modified_ts)
                    
                    # SetFileTime(handle, lpCreationTime, lpLastAccessTime, lpLastWriteTime)
                    kernel32.SetFileTime(
                        h,
                        ctypes.byref(ft_created) if ft_created else None,
                        None,  # Access time
                        ctypes.byref(ft_modified) if ft_modified else None
                    )
                    kernel32.CloseHandle(h)
                    print("[set_file_dates] Windows timestamps updated successfully")
                else:
                    print(f"[set_file_dates] Failed to open file handle, error: {ctypes.get_last_error()}")
                    
        except Exception as e:
            print(f"[set_file_dates] Error: {e}")
            import traceback
            traceback.print_exc()

